from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Alignment
from openpyxl import Workbook, load_workbook

file = 'folder/filename.xlsx'

writer = pd.ExcelWriter(file, engine = 'xlsxwriter')
accr_count_m.to_excel(writer, sheet_name = 'sheet_name', index = False, startrow = 2)
fac_no_accr_filtered.to_excel(writer, sheet_name = 'sheet_name2', index = False, startrow = 2)
data.to_excel(writer, sheet_name = 'sheet_name3', index = False, startrow = 2)

writer.save()
writer.close()

message1 = """This is a header that you can use in the upper part of your excel sheet to outline comments about the data set.
"""

message2 = """
Another header
"""
message3 = """
Another header
"""


wb = load_workbook(file)

sheet1 = wb['sheet_name'] 
sheet1['A1'] = message1
sheet1.merge_cells('A1:N1')
sheet1['A1'].font = Font(size=11, color='FF0013', bold=True, italic=False)
sheet1.row_dimensions[1].height = 170
sheet1['A1'].alignment = Alignment(wrapText=True)

sheet2 = wb['sheet_name2'] 
sheet2['A1'] = message2
sheet2.merge_cells('A1:N1')
sheet2['A1'].font = Font(size=11, color='FF0013', bold=True, italic=False)
sheet2.row_dimensions[1].height = 40
sheet2['A1'].alignment = Alignment(wrapText=True)

sheet3 = wb['sheet_name3'] 
sheet3['A1'] = message3
sheet3.merge_cells('A1:N1')
sheet3['A1'].font = Font(size=11, color='FF0013', bold=True, italic=False)
sheet3.row_dimensions[1].height = 40
sheet3['A1'].alignment = Alignment(wrapText=True)

wb.save(file)
